from __future__ import annotations

import argparse
import csv
from dataclasses import dataclass
from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation
import json
from pathlib import Path
from typing import Iterable


REQUIRED_COLUMNS = ("date", "category", "item", "amount")
FORMULA_PREFIXES = ("=", "+", "-", "@")


class InputContractError(ValueError):
    pass


@dataclass(frozen=True)
class SaleRecord:
    row_number: int
    date: date
    category: str
    item: str
    amount: Decimal


@dataclass(frozen=True)
class ValidationIssue:
    row_number: int
    reason: str
    raw: dict[str, str]


def safe_excel_text(value: str) -> str:
    if value.startswith(FORMULA_PREFIXES):
        return "'" + value
    return value


def normalize_raw(row: dict) -> dict[str, str]:
    normalized = {}
    for key, value in row.items():
        normalized["<extra>" if key is None else str(key)] = (
            "" if value is None else str(value)
        )
    return normalized


def parse_row(row_number: int, row: dict) -> SaleRecord:
    errors = []
    if row.get(None):
        errors.append("헤더보다 많은 열")

    raw_date = (row.get("date") or "").strip()
    try:
        parsed_date = date.fromisoformat(raw_date)
    except ValueError:
        errors.append("date는 YYYY-MM-DD 형식이어야 함")
        parsed_date = date.min

    category = (row.get("category") or "").strip()
    item = (row.get("item") or "").strip()
    if not category:
        errors.append("category가 비어 있음")
    if not item:
        errors.append("item이 비어 있음")

    raw_amount = (row.get("amount") or "").strip()
    try:
        amount = Decimal(raw_amount)
        if not amount.is_finite() or amount < 0:
            raise InvalidOperation
    except (InvalidOperation, ValueError):
        errors.append("amount는 0 이상의 유한한 숫자여야 함")
        amount = Decimal(0)

    if errors:
        raise InputContractError("; ".join(errors))
    return SaleRecord(row_number, parsed_date, category, item, amount)


def load_sales(input_path: Path) -> tuple[list[SaleRecord], list[ValidationIssue]]:
    path = input_path.expanduser().resolve()
    records = []
    issues = []
    try:
        stream = path.open("r", encoding="utf-8-sig", newline="")
    except OSError as exc:
        raise InputContractError(f"CSV를 읽을 수 없습니다: {path}") from exc

    with stream:
        reader = csv.DictReader(stream)
        if reader.fieldnames is None:
            raise InputContractError("CSV 헤더가 없습니다.")
        fieldnames = [name.strip() for name in reader.fieldnames if name is not None]
        if not all(fieldnames) or len(fieldnames) != len(set(fieldnames)):
            raise InputContractError("CSV 헤더가 비어 있거나 중복되었습니다.")
        missing = [name for name in REQUIRED_COLUMNS if name not in fieldnames]
        if missing:
            raise InputContractError(f"필수 열이 없습니다: {', '.join(missing)}")
        reader.fieldnames = fieldnames

        for row_number, row in enumerate(reader, start=2):
            raw = normalize_raw(row)
            try:
                records.append(parse_row(row_number, row))
            except InputContractError as exc:
                issues.append(ValidationIssue(row_number, str(exc), raw))

    return records, issues


def category_totals(records: Iterable[SaleRecord]) -> dict[str, Decimal]:
    totals: dict[str, Decimal] = {}
    for record in records:
        totals[record.category] = totals.get(record.category, Decimal(0)) + record.amount
    return dict(sorted(totals.items(), key=lambda item: item[0].lower()))


def style_header(cells) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill

    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in cells:
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")


def set_widths(sheet, widths: dict[str, float]) -> None:
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width


def build_workbook(records: list[SaleRecord], issues: list[ValidationIssue]):
    try:
        from openpyxl import Workbook
        from openpyxl.chart import BarChart, Reference
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as exc:
        raise RuntimeError(
            "openpyxl이 필요합니다. "
            "python -m pip install -r requirements-automate.txt를 실행하세요."
        ) from exc

    workbook = Workbook()
    summary = workbook.active
    summary.title = "요약"
    detail = workbook.create_sheet("정상 데이터")
    errors = workbook.create_sheet("검증 오류")

    summary.merge_cells("A1:D1")
    summary["A1"] = "CSV 매출 자동화 보고서"
    summary["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    summary["A1"].fill = PatternFill("solid", fgColor="17365D")
    summary["A1"].alignment = Alignment(horizontal="center")
    summary["A2"] = "생성 시각(UTC)"
    # Excel 날짜·시간 값에는 시간대 객체가 유지되지 않는다. 인접한 레이블에
    # 시간대 정보가 없는 이 값이 UTC 기준임을 명시한다.
    summary["B2"] = datetime.now(timezone.utc).replace(
        microsecond=0,
        tzinfo=None,
    )
    summary["B2"].number_format = "yyyy-mm-dd hh:mm:ss"

    summary["A4"] = "정상 건수"
    summary["B4"] = "=COUNTA('정상 데이터'!A:A)-1"
    summary["A5"] = "검증 오류 건수"
    summary["B5"] = "=COUNTA('검증 오류'!A:A)-1"
    summary["A6"] = "정상 데이터 총액"
    summary["B6"] = "=SUM('정상 데이터'!D:D)"
    summary["B6"].number_format = "#,##0.00"

    summary["A9"] = "분류"
    summary["B9"] = "합계"
    style_header(summary[9])
    for index, category in enumerate(category_totals(records), start=10):
        summary.cell(index, 1, safe_excel_text(category))
        summary.cell(
            index,
            2,
            f'=SUMIF(\'정상 데이터\'!B:B,A{index},\'정상 데이터\'!D:D)',
        )
        summary.cell(index, 2).number_format = "#,##0.00"

    detail.append(["날짜", "분류", "항목", "금액", "원본 행"])
    style_header(detail[1])
    for record in records:
        detail.append(
            [
                record.date,
                safe_excel_text(record.category),
                safe_excel_text(record.item),
                float(record.amount),
                record.row_number,
            ]
        )
    for cell in detail["A"][1:]:
        cell.number_format = "yyyy-mm-dd"
    for cell in detail["D"][1:]:
        cell.number_format = "#,##0.00"
    detail.freeze_panes = "A2"
    detail.auto_filter.ref = detail.dimensions

    errors.append(["원본 행", "오류 사유", "원본 값(JSON)"])
    style_header(errors[1])
    for issue in issues:
        errors.append(
            [
                issue.row_number,
                safe_excel_text(issue.reason),
                json.dumps(issue.raw, ensure_ascii=False, sort_keys=True),
            ]
        )
    errors.freeze_panes = "A2"
    errors.auto_filter.ref = errors.dimensions

    if category_totals(records):
        chart = BarChart()
        chart.title = "분류별 합계"
        chart.y_axis.title = "금액"
        chart.x_axis.title = "분류"
        last_row = 9 + len(category_totals(records))
        chart.add_data(Reference(summary, min_col=2, min_row=9, max_row=last_row), titles_from_data=True)
        chart.set_categories(Reference(summary, min_col=1, min_row=10, max_row=last_row))
        chart.height = 7
        chart.width = 12
        summary.add_chart(chart, "D4")

    set_widths(summary, {"A": 24, "B": 20, "C": 3, "D": 16})
    set_widths(detail, {"A": 14, "B": 20, "C": 30, "D": 16, "E": 12})
    set_widths(errors, {"A": 12, "B": 48, "C": 80})
    summary.freeze_panes = "A4"

    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"
    return workbook


def save_workbook(workbook, output_path: Path) -> Path:
    output = output_path.expanduser().resolve()
    if output.suffix.lower() != ".xlsx":
        raise ValueError("출력 확장자는 .xlsx여야 합니다.")
    output.parent.mkdir(parents=True, exist_ok=True)
    temporary = output.with_name(output.stem + ".tmp.xlsx")
    workbook.save(temporary)
    temporary.replace(output)
    return output


def verify_workbook(output_path: Path, expected_records: int, expected_issues: int) -> None:
    from openpyxl import load_workbook

    workbook = load_workbook(output_path, data_only=False, read_only=True)
    expected_sheets = ["요약", "정상 데이터", "검증 오류"]
    if workbook.sheetnames != expected_sheets:
        raise ValueError(f"시트 구조가 잘못되었습니다: {workbook.sheetnames}")
    if workbook["정상 데이터"].max_row - 1 != expected_records:
        raise ValueError("정상 데이터 행 수가 일치하지 않습니다.")
    if workbook["검증 오류"].max_row - 1 != expected_issues:
        raise ValueError("검증 오류 행 수가 일치하지 않습니다.")
    if workbook["요약"]["B4"].value != "=COUNTA('정상 데이터'!A:A)-1":
        raise ValueError("요약 검증 수식이 없습니다.")
    workbook.close()


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="CSV를 검증하고 Excel 요약 보고서를 생성합니다.",
    )
    parser.add_argument("input", type=Path)
    parser.add_argument("--output", type=Path, default=Path("sales-report.xlsx"))
    parser.add_argument(
        "--allow-invalid",
        action="store_true",
        help="검증 오류가 있어도 종료 코드 0을 사용합니다.",
    )
    return parser


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    try:
        records, issues = load_sales(args.input)
        workbook = build_workbook(records, issues)
        output = save_workbook(workbook, args.output)
        verify_workbook(output, len(records), len(issues))
    except (InputContractError, OSError, RuntimeError, ValueError) as exc:
        print(f"보고서 생성 실패: {exc}")
        return 1

    print(f"보고서 생성 완료: {output}")
    print(f"정상 {len(records)}건, 검증 오류 {len(issues)}건")
    if issues and not args.allow_invalid:
        print("검증 오류가 있어 종료 코드 2를 반환합니다.")
        return 2
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
